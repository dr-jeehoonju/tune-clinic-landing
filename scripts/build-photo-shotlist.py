#!/usr/bin/env python3
"""
Build the 2026-04-27 clinic photo shotlist.

Outputs:
    docs/photo-shotlist-2026-04-27.xlsx  (3 sheets: Shotlist, Summary by Page, Shoot Day Plan)
    docs/photo-shotlist-2026-04-27.md     (markdown table + intro)

The row dataset below is the source of truth. Edit ROWS, BLOCKS and PAGE_NOTES
and rerun this script to regenerate.
"""
from __future__ import annotations

import os
from collections import OrderedDict, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
OUT_XLSX = os.path.join(REPO_ROOT, "docs", "photo-shotlist-2026-04-27.xlsx")
OUT_MD = os.path.join(REPO_ROOT, "docs", "photo-shotlist-2026-04-27.md")

# ---------------------------------------------------------------------------
# Column schema
# ---------------------------------------------------------------------------
COLUMNS = [
    ("id", 10),
    ("priority", 10),
    ("page", 22),
    ("page_file", 38),
    ("section", 30),
    ("current_state", 18),
    ("photo_subject", 56),
    ("composition", 16),
    ("orientation", 18),
    ("min_resolution", 16),
    ("mood_lighting", 28),
    ("must_include", 44),
    ("must_avoid", 40),
    ("consent_required", 22),
    ("suggested_filename", 50),
    ("count", 14),
    ("notes", 50),
]

WRAP_COLS = {"photo_subject", "must_include", "must_avoid", "notes", "section"}

PRIORITY_FILL = {
    "P0": PatternFill("solid", fgColor="FECACA"),
    "P1": PatternFill("solid", fgColor="FED7AA"),
    "P2": PatternFill("solid", fgColor="E2E8F0"),
}

# ---------------------------------------------------------------------------
# Shotlist rows
#
# Field order matches COLUMNS. Use empty string for n/a.
# Conventions:
#   - filenames live under src/img/photos/{slug}/...
#   - .webp final delivery
#   - "current_state" = what the page renders today (text-only / video poster /
#     /main.jpeg fallback / etc.)
# ---------------------------------------------------------------------------

# Reusable shorthand
LANDSCAPE_HERO = "landscape (16:9)"
LANDSCAPE_3_2 = "landscape (3:2)"
SQUARE = "square (1:1)"
PORTRAIT = "portrait (4:5)"

DEF_HERO_RES = "2400x1350"
DEF_LARGE_RES = "2000x1333"
DEF_SQUARE_RES = "1600x1600"
DEF_PORTRAIT_RES = "1600x2000"
DEF_THUMB_RES = "1200x1200"

ROWS: list[dict] = []


def add(**kw):
    ROWS.append(kw)


# ============================================================
# P0 — critical for 5/6 ad launch
# ============================================================

# Home (/)
add(
    priority="P0",
    page="/",
    page_file="src/fragments/en/index.html",
    section="Hero (poster fallback for /hero-video.mp4)",
    current_state="video w/ /main.jpeg poster",
    photo_subject="Wide cinematic shot of Tune Clinic Apgujeong building exterior at golden hour with subtle gold accent lighting; signage legible. Used as <video poster> when video fails or is suppressed.",
    composition="wide",
    orientation=LANDSCAPE_HERO,
    min_resolution="2880x1620",
    mood_lighting="warm golden hour, premium",
    must_include="Tune Clinic signage, 5th-floor building facade, clean storefront",
    must_avoid="passersby faces, license plates, competing clinic signs in same frame",
    consent_required="No",
    suggested_filename="src/img/photos/home/hero-clinic-exterior-golden-hour.webp",
    count="3 angles, 1 final",
    notes="Also produce a square 1:1 crop for OG; align horizon with rule-of-thirds so headline overlay (left-bottom) stays uncluttered.",
)
add(
    priority="P0",
    page="/",
    page_file="src/fragments/en/index.html",
    section="Hero (alternate interior poster)",
    current_state="video w/ /main.jpeg poster",
    photo_subject="Signature interior wide of consultation room or main hallway with gold accent + warm wood — composed dark on edges so white serif headline reads.",
    composition="wide",
    orientation=LANDSCAPE_HERO,
    min_resolution="2880x1620",
    mood_lighting="soft moody premium, warm tungsten",
    must_include="clinic interior detail (gold trim, wood, plants), uncluttered surface",
    must_avoid="staff faces (unless released), patient personal items, glare on glass",
    consent_required="No",
    suggested_filename="src/img/photos/home/hero-interior-signature.webp",
    count="2 angles",
    notes="Optional secondary poster; leave 40% negative space on left for headline.",
)
add(
    priority="P0",
    page="/",
    page_file="src/fragments/en/index.html",
    section="Why Apgujeong (trust strip, line 124)",
    current_state="text only",
    photo_subject="Apgujeong-rodeo street view from clinic's 5th-floor window OR from sidewalk looking up at building — establishes the district context.",
    composition="wide",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="bright daylight, slightly desaturated premium",
    must_include="recognizable Apgujeong streetscape (boutiques, tree-lined road)",
    must_avoid="identifiable pedestrians, brand logos of luxury houses without permission",
    consent_required="No",
    suggested_filename="src/img/photos/home/why-apgujeong-streetview.webp",
    count="2",
    notes="Reuses on /apgujeong-aesthetic-clinic-for-foreign-patients.html and /international-patients-guide.html.",
)
add(
    priority="P0",
    page="/",
    page_file="src/fragments/en/index.html",
    section="Inside the Clinic (line 1256, dark CTA section)",
    current_state="text only",
    photo_subject="Reception desk with concierge member greeting (back-of-head only or signed-release talent), warm lamp, gold accent visible.",
    composition="medium",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="warm tungsten, evening premium",
    must_include="reception counter, branded signage subtle, fresh florals",
    must_avoid="visible patient face, prices on monitor, paperwork with names",
    consent_required="Yes (staff release)",
    suggested_filename="src/img/photos/home/inside-clinic-reception.webp",
    count="3 angles",
    notes="This block is currently 100% text on dark; image is high-impact.",
)
add(
    priority="P0",
    page="/",
    page_file="src/fragments/en/index.html",
    section="Visit Us (line 1335 - location/map block)",
    current_state="text only",
    photo_subject="5th-floor elevator landing or floor signage immediately outside clinic door — orientation cue for arriving patients.",
    composition="medium",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="clinical bright neutral",
    must_include="floor number visible, clinic logo on door/wall, polished floor",
    must_avoid="other tenants' signage in same frame, cleaning carts",
    consent_required="No",
    suggested_filename="src/img/photos/home/visit-us-5f-arrival.webp",
    count="2",
    notes="Pairs naturally next to the address text + map embed.",
)

# Booking
add(
    priority="P0",
    page="/booking.html",
    page_file="src/fragments/en/booking.html",
    section="Hero (line 4)",
    current_state="text only, gold underline only",
    photo_subject="Calm consultation desk overhead — leather notebook, clinic-branded card, fountain pen, single fresh stem in vase, soft hands of physician (Dr. Cha or Dr. Ju) writing a plan.",
    composition="overhead",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="soft daylight from left, premium calm",
    must_include="hand in white-coat cuff, branded card, planning notebook",
    must_avoid="real patient name on card, phone screens with private content",
    consent_required="Yes (physician release)",
    suggested_filename="src/img/photos/booking/hero-consultation-desk-overhead.webp",
    count="3 variants",
    notes="Doubles as 'plan with the doctor' visual; reuse across decision-protection and design-method.",
)
add(
    priority="P0",
    page="/booking.html",
    page_file="src/fragments/en/booking.html",
    section="Booking confirmation success state (line 200)",
    current_state="text only with green checkmark",
    photo_subject="Soft macro of fresh florals + 'Welcome' note card on reception counter — used as gentle reassurance image after booking.",
    composition="close-up",
    orientation=SQUARE,
    min_resolution=DEF_SQUARE_RES,
    mood_lighting="soft daylight, airy",
    must_include="florals (white/cream), branded card or envelope",
    must_avoid="any text that could date the image, patient names",
    consent_required="No",
    suggested_filename="src/img/photos/booking/success-welcome-florals.webp",
    count="2",
    notes="Optional but elevates UX after submit.",
)

# Metacell Protocol
add(
    priority="P0",
    page="/metacell-protocol.html",
    page_file="src/fragments/en/metacell-protocol.html",
    section="Hero (line 35) — flagship LP",
    current_state="text only on dark gradient",
    photo_subject="Physician hands (gloved) holding a labeled PRP centrifuge tube against a soft bokeh of treatment-room lighting; alternate take with PBM device head in foreground.",
    composition="close-up",
    orientation=LANDSCAPE_HERO,
    min_resolution="2880x1620",
    mood_lighting="soft moody premium, blue-warm split",
    must_include="gloved hands, PRP tube OR PBM device, white-coat cuff",
    must_avoid="any 'stem cell' wording on labels (regulatory), patient body parts",
    consent_required="Yes (physician release)",
    suggested_filename="src/img/photos/metacell/hero-prp-pbm-hands.webp",
    count="4 variants",
    notes="Critical for ad launch. Compose dark-left so white headline reads.",
)
add(
    priority="P0",
    page="/metacell-protocol.html",
    page_file="src/fragments/en/metacell-protocol.html",
    section="Autologous regenerative explainer (line 80)",
    current_state="text only",
    photo_subject="PRP centrifuge machine on stainless tray with prepared tubes mid-process — clean clinical scene.",
    composition="medium",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="clinical bright neutral",
    must_include="centrifuge, labeled tubes (Tune Clinic label OK), gloved hand entering frame",
    must_avoid="cluttered countertop, expired-looking labels",
    consent_required="No",
    suggested_filename="src/img/photos/metacell/equipment-prp-centrifuge.webp",
    count="3 angles",
    notes="Equipment close-up reused on collagen-builder + skin-boosters pages.",
)
add(
    priority="P0",
    page="/metacell-protocol.html",
    page_file="src/fragments/en/metacell-protocol.html",
    section="Three illustrative builds (line 109)",
    current_state="text only, three card grid",
    photo_subject="Three matched square crops: (a) Ultherapy device handpiece, (b) Thermage tip on machine, (c) PBM device head — same lighting setup so cards align.",
    composition="close-up",
    orientation=SQUARE,
    min_resolution=DEF_SQUARE_RES,
    mood_lighting="clinical bright neutral with single warm rim light",
    must_include="device branding faintly visible, dust-free surfaces",
    must_avoid="serial numbers, mismatched lighting between the three",
    consent_required="No",
    suggested_filename="src/img/photos/metacell/build-{ultherapy|thermage|pbm}-card.webp",
    count="3 (one per build)",
    notes="Shoot all three back-to-back with same lighting to keep card grid consistent.",
)

# Doctor portraits — used everywhere
add(
    priority="P0",
    page="(global)",
    page_file="src/fragments/en/index.html (Medical Board, line 893)",
    section="Dr. Cha — formal portrait",
    current_state="initials avatar / text only",
    photo_subject="Dr. Seung Yeon Cha, Representative Director — half-body in white coat against warm neutral wall, calm confident expression looking just off-camera.",
    composition="portrait",
    orientation=PORTRAIT,
    min_resolution=DEF_PORTRAIT_RES,
    mood_lighting="soft window daylight + bounce, premium clinical",
    must_include="white coat with subtle Tune branding, name pin if available, neutral background",
    must_avoid="strong shadows, busy background, distracting jewelry",
    consent_required="Yes (physician release)",
    suggested_filename="src/img/photos/doctors/dr-cha-portrait-formal.webp",
    count="3 variants (head-shot, half-body, full)",
    notes="Most-used asset on the site. Capture matching landscape crop too (3:2).",
)
add(
    priority="P0",
    page="(global)",
    page_file="src/fragments/en/index.html",
    section="Dr. Cha — environmental",
    current_state="text only",
    photo_subject="Dr. Cha at consultation desk, half-body, looking at patient chart / tablet (no patient face visible) — body angled 30deg from camera.",
    composition="medium",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="warm natural daylight from window-left",
    must_include="consultation chair, planning notebook, gold accent",
    must_avoid="real patient identifiers on chart, screen content with names",
    consent_required="Yes (physician release)",
    suggested_filename="src/img/photos/doctors/dr-cha-consultation.webp",
    count="3",
    notes="Used on /booking, /design-method, /decision-protection, /metacell-protocol footer CTAs.",
)
add(
    priority="P0",
    page="(global)",
    page_file="src/fragments/en/index.html",
    section="Dr. Ju — formal portrait",
    current_state="initials avatar / text only",
    photo_subject="Dr. Jee Hoon Ju, International Director — half-body in white coat, friendly approachable expression, slight smile, looking at camera.",
    composition="portrait",
    orientation=PORTRAIT,
    min_resolution=DEF_PORTRAIT_RES,
    mood_lighting="soft window daylight + bounce, premium clinical",
    must_include="white coat, English-speaking-lead-friendly demeanor, neutral background",
    must_avoid="overly serious/clinical mood — should read 'approachable' for international patients",
    consent_required="Yes (physician release)",
    suggested_filename="src/img/photos/doctors/dr-ju-portrait-formal.webp",
    count="3 variants",
    notes="Pair with matching landscape crop for use as English-language LP author byline.",
)
add(
    priority="P0",
    page="(global)",
    page_file="src/fragments/en/index.html",
    section="Dr. Ju — environmental (English consult)",
    current_state="text only",
    photo_subject="Dr. Ju at consultation table with iPad showing planning visual, gesturing in conversation across desk (back-of-head of stand-in patient OK).",
    composition="medium",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="warm daylight + warm tungsten fill",
    must_include="iPad with planning visual (placeholder), Tune card, two chairs",
    must_avoid="real patient face, real chart data on iPad",
    consent_required="Yes (physician + stand-in release)",
    suggested_filename="src/img/photos/doctors/dr-ju-international-consult.webp",
    count="3",
    notes="Anchors English-speaking concierge story across booking + international guide.",
)

# Building / location proof
add(
    priority="P0",
    page="(global)",
    page_file="src/fragments/en/apgujeong-aesthetic-clinic-for-foreign-patients.html",
    section="Building exterior wide",
    current_state="text only",
    photo_subject="Wide street-level shot of 868 Nonhyeon-ro building, full facade with 5th-floor location identifiable; pedestrian-free if possible.",
    composition="wide",
    orientation=LANDSCAPE_HERO,
    min_resolution="2880x1620",
    mood_lighting="bright midday daylight (sunny window 11am)",
    must_include="full building, street-level context, address number subtle",
    must_avoid="people in the frame, parked cars dominating",
    consent_required="No",
    suggested_filename="src/img/photos/location/building-exterior-wide.webp",
    count="3 (different focal lengths)",
    notes="Reused as OG image candidate for apgujeong page.",
)
add(
    priority="P0",
    page="(global)",
    page_file="src/fragments/en/apgujeong-aesthetic-clinic-for-foreign-patients.html",
    section="5th-floor reception (location proof)",
    current_state="text only",
    photo_subject="Reception/waiting lounge wide — clean waiting seats, tea station, branded panel, plant.",
    composition="wide",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="warm daylight + accent lighting",
    must_include="branded wall, comfortable seating, tea/water station, fresh flowers",
    must_avoid="patients in waiting, magazines with names",
    consent_required="No",
    suggested_filename="src/img/photos/location/reception-waiting-wide.webp",
    count="3 angles",
    notes="Shoot before clinic opens for empty waiting area.",
)
add(
    priority="P0",
    page="(global)",
    page_file="src/fragments/en/apgujeong-aesthetic-clinic-for-foreign-patients.html",
    section="Apgujeong-rodeo street context",
    current_state="text only",
    photo_subject="Apgujeong-rodeo street scene mid-block — boutique storefronts, tree canopy, leaves, premium district vibe.",
    composition="wide",
    orientation=LANDSCAPE_HERO,
    min_resolution="2880x1620",
    mood_lighting="bright daylight, slightly desaturated",
    must_include="recognizable Apgujeong-rodeo streetscape",
    must_avoid="identifiable individual pedestrians, competitor clinics",
    consent_required="No",
    suggested_filename="src/img/photos/location/apgujeong-rodeo-street.webp",
    count="4 (variations)",
    notes="Multi-page reuse — home, apgujeong page, international guide, blog hero pool.",
)

# ============================================================
# P1 — within 2 weeks
# ============================================================

# Procedure LP heroes
add(
    priority="P1",
    page="/signature-lifting.html",
    page_file="src/fragments/en/signature-lifting.html",
    section="Hero (line 37)",
    current_state="text only on dark",
    photo_subject="Ultherapy device with handpiece in physician's hand, mid-treatment angle (no patient face) — handpiece resting on neutral pad with gel.",
    composition="medium",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="soft moody premium, dark backdrop",
    must_include="Ultherapy branding visible but subtle, gloved hand",
    must_avoid="patient face, procedure-in-progress with skin visible",
    consent_required="Yes (physician release)",
    suggested_filename="src/img/photos/signature-lifting/hero-ultherapy-handpiece.webp",
    count="3",
    notes="Compose dark-top so white headline reads.",
)
add(
    priority="P1",
    page="/structural-reset.html",
    page_file="src/fragments/en/structural-reset.html",
    section="Hero (line 35)",
    current_state="text only on dark",
    photo_subject="Thermage FLX device with tip and CPT cable visible — sense of premium engineering.",
    composition="medium",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="soft moody premium",
    must_include="Thermage branding subtle, hand entering frame",
    must_avoid="patient body, used disposables in frame",
    consent_required="No",
    suggested_filename="src/img/photos/structural-reset/hero-thermage-tip.webp",
    count="3",
    notes="VIP slot LP — emphasize premium feel.",
)
add(
    priority="P1",
    page="/collagen-builder.html",
    page_file="src/fragments/en/collagen-builder.html",
    section="Hero (line 35)",
    current_state="text only on dark",
    photo_subject="Tray of skin-booster vials (Rejuran / Juvelook) being arranged by gloved hand — clinical premium.",
    composition="overhead",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="soft moody premium with single rim light",
    must_include="vials with manufacturer branding, sterile tray, gloved hand",
    must_avoid="opened needle caps in unsafe orientation, patient",
    consent_required="No",
    suggested_filename="src/img/photos/collagen-builder/hero-skin-booster-tray.webp",
    count="3",
    notes="Shows 'design before injection' tone.",
)
add(
    priority="P1",
    page="/filler-chamaka-se.html",
    page_file="src/fragments/en/filler-chamaka-se.html",
    section="Hero (line 35)",
    current_state="text only on dark",
    photo_subject="Physician holding face-mapping marker over a face-anatomy diagram or 3D mannequin head — design intent.",
    composition="close-up",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="soft moody premium",
    must_include="anatomy mannequin or printed face map, marker, gloved hand",
    must_avoid="photo of real patient, comparing brands",
    consent_required="Yes (physician release)",
    suggested_filename="src/img/photos/filler-chamaka-se/hero-design-mapping.webp",
    count="3",
    notes="Doubles for 'Designed for Harmony' section.",
)
add(
    priority="P1",
    page="/decision-protection.html",
    page_file="src/fragments/en/decision-protection.html",
    section="Hero (line 14)",
    current_state="text only",
    photo_subject="Calm consultation desk with iPad showing planning visual, paper plan, glass of water, no people — invitation to think.",
    composition="overhead",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="soft daylight, airy minimalist",
    must_include="iPad with placeholder plan visual, printed plan card, fresh stem, branded notebook",
    must_avoid="phones, names, real patient data",
    consent_required="No",
    suggested_filename="src/img/photos/decision-protection/hero-planning-desk.webp",
    count="3",
    notes="Reuse for /metacell-protocol footer CTA and /design-method intro.",
)
add(
    priority="P1",
    page="/design-method.html",
    page_file="src/fragments/en/design-method.html",
    section="Hero (line 37)",
    current_state="text only",
    photo_subject="Physician hand drawing facial vector lines on a printed face anatomy sheet — anatomical design feel.",
    composition="close-up",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="warm daylight, focused desk lamp",
    must_include="anatomy diagram, fineliner or marker, hand only",
    must_avoid="patient photo on the desk, brand logos of devices",
    consent_required="Yes (physician release, hand-only OK)",
    suggested_filename="src/img/photos/design-method/hero-vector-mapping.webp",
    count="3",
    notes="Captures 'Anatomical Design Workflow' theme.",
)
add(
    priority="P1",
    page="/design-method.html",
    page_file="src/fragments/en/design-method.html",
    section="Anatomical Design Workflow steps (line 130)",
    current_state="text only, six step cards",
    photo_subject="Series of six matched close-ups: assessment notes, planning iPad, vector overlay sheet, structural blueprint, micro-detail tools, depth-safety calibration card.",
    composition="close-up",
    orientation=SQUARE,
    min_resolution=DEF_SQUARE_RES,
    mood_lighting="warm daylight, consistent across all six",
    must_include="consistent surface (same desk), one hand max per frame",
    must_avoid="patient identifiers, mismatched lighting",
    consent_required="Yes (physician release for hands)",
    suggested_filename="src/img/photos/design-method/workflow-step-{01-06}.webp",
    count="6 cards (matched set)",
    notes="Shoot all six in one sitting for visual consistency.",
)
add(
    priority="P1",
    page="/apgujeong-aesthetic-clinic-for-foreign-patients.html",
    page_file="src/fragments/en/apgujeong-aesthetic-clinic-for-foreign-patients.html",
    section="Hero (line 17)",
    current_state="text only",
    photo_subject="Combination split: building facade left + cropped Apgujeong-rodeo skyline right, OR a single hero of the 5F approach with district context through window.",
    composition="wide",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="bright daylight, premium",
    must_include="recognizable Apgujeong context + clinic identifier",
    must_avoid="competitor signs",
    consent_required="No",
    suggested_filename="src/img/photos/apgujeong/hero-district-clinic.webp",
    count="2",
    notes="Can compose from existing P0 location shots if shoot day is tight.",
)
add(
    priority="P1",
    page="/aesthetic-treatment-faq-for-foreign-patients.html",
    page_file="src/fragments/en/aesthetic-treatment-faq-for-foreign-patients.html",
    section="Hero (line 14)",
    current_state="text only",
    photo_subject="Front-desk concierge at reception with iPad in English UI, calmly answering question (back of patient head visible only).",
    composition="medium",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="warm daylight",
    must_include="concierge in uniform/blazer, iPad in English UI, branded counter",
    must_avoid="real patient face, real iPad data",
    consent_required="Yes (staff release; stand-in patient release)",
    suggested_filename="src/img/photos/faq/hero-concierge-english-support.webp",
    count="3",
    notes="Supports 'English support is real' trust.",
)
add(
    priority="P1",
    page="/guides.html",
    page_file="src/fragments/en/guides.html",
    section="Hero (line 31)",
    current_state="text only",
    photo_subject="Printed Tune Clinic guides / brochures fanned on consultation desk with reading glasses + warm coffee — pre-arrival planning vibe.",
    composition="overhead",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="warm daylight, airy",
    must_include="branded guide covers (placeholder OK), reading aid, warm cup",
    must_avoid="competitor brochures in frame",
    consent_required="No",
    suggested_filename="src/img/photos/guides/hero-printed-guides.webp",
    count="2",
    notes="Pair with smaller marquee crop for /guides sub-pages.",
)
add(
    priority="P1",
    page="/korean-lifting-guide.html",
    page_file="src/fragments/en/korean-lifting-guide.html",
    section="Hero (line 16) + device cards (line 39)",
    current_state="text only",
    photo_subject="Lineup overhead of three lifting devices' branded heads (Ultherapy, Thermage FLX, Oligio/ONDA) on neutral pad — visual encyclopedia.",
    composition="overhead",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="clinical bright neutral",
    must_include="three handpieces clearly distinguishable, clean pad",
    must_avoid="duct tape, dust, mismatched lighting",
    consent_required="No",
    suggested_filename="src/img/photos/korean-lifting-guide/hero-devices-lineup.webp",
    count="2 (overhead + 30deg)",
    notes="Each device also gets its own square crop for the four cards.",
)
add(
    priority="P1",
    page="/ultherapy-vs-thermage.html",
    page_file="src/fragments/en/ultherapy-vs-thermage.html",
    section="Hero (line 14)",
    current_state="text only",
    photo_subject="Ultherapy and Thermage FLX devices side by side in same frame, separated by neutral panel — no winner implied.",
    composition="wide",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="clinical bright neutral with soft side fill",
    must_include="both device fronts, branding subtle but visible",
    must_avoid="any visual ranking (e.g. one elevated)",
    consent_required="No",
    suggested_filename="src/img/photos/ultherapy-vs-thermage/hero-side-by-side.webp",
    count="2",
    notes="Critical to keep visually balanced — equal framing.",
)
add(
    priority="P1",
    page="/dermal-fillers-in-korea.html",
    page_file="src/fragments/en/dermal-fillers-in-korea.html",
    section="Hero (line 15) + brand cards (line 36)",
    current_state="text only",
    photo_subject="Refrigerated tray of HA filler boxes (Juvederm, Restylane, Belotero, Radiesse) arranged neatly — 'real clinical flow' theme.",
    composition="overhead",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="clinical bright neutral, slight cool tone",
    must_include="visible HA filler brand boxes (not expired), gloved hand",
    must_avoid="open/used syringes, no brand displayed dominantly larger than others",
    consent_required="No",
    suggested_filename="src/img/photos/dermal-fillers/hero-filler-tray.webp",
    count="2 (full + tight crop)",
    notes="Compliance: equal-sized brand presence; do not imply endorsement.",
)
add(
    priority="P1",
    page="/skin-boosters-and-regenerative-treatments.html",
    page_file="src/fragments/en/skin-boosters-and-regenerative-treatments.html",
    section="Hero (line 14) + brand cards (line 35)",
    current_state="text only",
    photo_subject="Tray of skin-booster vials (Rejuran/PDRN, Juvelook, exosomes, Sculptra) with labels visible, neat clinical arrangement.",
    composition="overhead",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="clinical bright neutral",
    must_include="four product categories visible, sterile tray",
    must_avoid="any brand visually dominant, opened sharps",
    consent_required="No",
    suggested_filename="src/img/photos/skin-boosters/hero-vial-tray.webp",
    count="2",
    notes="Pair with single-product close-ups for the four cards.",
)
add(
    priority="P1",
    page="/hair-loss-and-scalp-regeneration-in-korea.html",
    page_file="src/fragments/en/hair-loss-and-scalp-regeneration-in-korea.html",
    section="Hero (line 14)",
    current_state="text only",
    photo_subject="Scalp-treatment device head over a neutral mannequin scalp model OR overhead of scalp-care vial set + applicator — no real patient.",
    composition="medium",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="clinical bright neutral",
    must_include="scalp device or labelled vials, gloved hand",
    must_avoid="visible scalp of real patient, hair loss before/after photos",
    consent_required="No",
    suggested_filename="src/img/photos/hair-loss/hero-scalp-treatment.webp",
    count="2",
    notes="Lower-traffic page; one solid hero is enough.",
)
add(
    priority="P1",
    page="/how-to-choose-the-right-lifting-treatment.html",
    page_file="src/fragments/en/how-to-choose-the-right-lifting-treatment.html",
    section="Hero (line 14) + indication cards (line 30)",
    current_state="text only",
    photo_subject="Physician's hand drawing four indication zones on printed face diagram (sagging, loose skin, contour collapse, travel timing).",
    composition="close-up",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="warm daylight",
    must_include="anatomy diagram with annotations, marker",
    must_avoid="patient photo, brand-name device",
    consent_required="Yes (physician release; hand-only OK)",
    suggested_filename="src/img/photos/how-to-choose/hero-indication-mapping.webp",
    count="2",
    notes="Same lighting as design-method workflow set.",
)
add(
    priority="P1",
    page="/international-patients-guide.html",
    page_file="src/fragments/en/international-patients-guide.html",
    section="Hero (line 19)",
    current_state="text only",
    photo_subject="Open passport, printed itinerary, branded planning card, AirPods on minimalist desk — 'planning your Seoul trip' overhead.",
    composition="overhead",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="warm daylight, airy",
    must_include="passport (cover only or generic), itinerary print, branded card",
    must_avoid="real passport photo page, real flight numbers/names",
    consent_required="No",
    suggested_filename="src/img/photos/international-guide/hero-trip-planning-desk.webp",
    count="2",
    notes="Pair with concierge image for FAQ page.",
)
add(
    priority="P1",
    page="/menu.html",
    page_file="src/fragments/en/menu.html",
    section="Hero (line 111)",
    current_state="text only on dark",
    photo_subject="Printed Chamaka-se menu booklet on premium paper, edge-lit on stone counter, single fresh stem.",
    composition="overhead",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="soft moody premium",
    must_include="Tune Clinic-branded menu (placeholder OK), stone surface",
    must_avoid="competitor pricing in shot",
    consent_required="No",
    suggested_filename="src/img/photos/menu/hero-printed-menu.webp",
    count="2",
    notes="Helps the price-list page feel less like a spreadsheet.",
)

# Equipment close-ups
add(
    priority="P1",
    page="(equipment library)",
    page_file="(reused across LPs)",
    section="Ultherapy device close-up",
    current_state="text only",
    photo_subject="Ultherapy device handpiece with Tune branding sticker subtle, mounted on machine.",
    composition="close-up",
    orientation=SQUARE,
    min_resolution=DEF_SQUARE_RES,
    mood_lighting="clinical bright neutral",
    must_include="Ultherapy branding, serial number obscured",
    must_avoid="dust, used gel residue",
    consent_required="No",
    suggested_filename="src/img/photos/equipment/ultherapy-handpiece.webp",
    count="3 angles",
    notes="Reused on signature-lifting, structural-reset, korean-lifting-guide, ultherapy-vs-thermage.",
)
add(
    priority="P1",
    page="(equipment library)",
    page_file="(reused across LPs)",
    section="Thermage FLX device close-up",
    current_state="text only",
    photo_subject="Thermage FLX tip with CPT cartridge attached, on machine; clean.",
    composition="close-up",
    orientation=SQUARE,
    min_resolution=DEF_SQUARE_RES,
    mood_lighting="clinical bright neutral",
    must_include="Thermage branding subtle, tip facing camera",
    must_avoid="dust, used cartridges visible",
    consent_required="No",
    suggested_filename="src/img/photos/equipment/thermage-flx-tip.webp",
    count="3 angles",
    notes="Reused on structural-reset, ultherapy-vs-thermage.",
)
add(
    priority="P1",
    page="(equipment library)",
    page_file="(reused across LPs)",
    section="ONDA Lifting device close-up",
    current_state="text only",
    photo_subject="ONDA microwave-lifting handpiece on machine — distinctive shape.",
    composition="close-up",
    orientation=SQUARE,
    min_resolution=DEF_SQUARE_RES,
    mood_lighting="clinical bright neutral",
    must_include="ONDA branding, clean handpiece",
    must_avoid="cluttered backdrop",
    consent_required="No",
    suggested_filename="src/img/photos/equipment/onda-lifting.webp",
    count="2",
    notes="Reused on korean-lifting-guide, signature-lifting.",
)
add(
    priority="P1",
    page="(equipment library)",
    page_file="(reused across LPs)",
    section="PBM (photobiomodulation) device",
    current_state="text only",
    photo_subject="PBM device head with red light visible at low intensity, on stand — minimalist.",
    composition="close-up",
    orientation=SQUARE,
    min_resolution=DEF_SQUARE_RES,
    mood_lighting="clinical bright neutral with red device glow",
    must_include="device head, soft red glow, no patient",
    must_avoid="overexposed light source",
    consent_required="No",
    suggested_filename="src/img/photos/equipment/pbm-device.webp",
    count="2",
    notes="Critical for /metacell-protocol cards.",
)
add(
    priority="P1",
    page="(equipment library)",
    page_file="(reused across LPs)",
    section="Juvelook syringe / Rejuran vials",
    current_state="text only",
    photo_subject="Single Juvelook syringe and Rejuran box arranged with sterile pad — premium product still life.",
    composition="close-up",
    orientation=SQUARE,
    min_resolution=DEF_SQUARE_RES,
    mood_lighting="soft daylight, clean white surface",
    must_include="manufacturer branding intact, lot label obscured",
    must_avoid="opened needle, expired packaging",
    consent_required="No",
    suggested_filename="src/img/photos/equipment/juvelook-rejuran-still.webp",
    count="3 (separate + together)",
    notes="Reused on collagen-builder, skin-boosters, blog hero pool.",
)

# Comfort / amenities
add(
    priority="P1",
    page="/structural-reset.html, /menu.html",
    page_file="src/fragments/en/menu.html (Comfort Experience, line 598)",
    section="IV sedation chair",
    current_state="text only",
    photo_subject="Empty IV-sedation recliner with warm blanket folded, side table with water — calm comfort.",
    composition="medium",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="warm daylight + warm tungsten",
    must_include="recliner, folded throw, side table, no patient",
    must_avoid="visible used IV bags, monitors with data",
    consent_required="No",
    suggested_filename="src/img/photos/comfort/iv-sedation-chair.webp",
    count="2",
    notes="High emotional payoff for premium audiences.",
)
add(
    priority="P1",
    page="/structural-reset.html, /menu.html",
    page_file="src/fragments/en/menu.html",
    section="Recovery lounge",
    current_state="text only",
    photo_subject="Recovery lounge wide — ambient lighting, soft seating, fresh tea station, quiet plant — no patient.",
    composition="wide",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="warm tungsten, soothing premium",
    must_include="ambient seating, tea station, plants",
    must_avoid="patients, used towels",
    consent_required="No",
    suggested_filename="src/img/photos/comfort/recovery-lounge.webp",
    count="3 angles",
    notes="Shoot pre-opening for empty space.",
)
add(
    priority="P1",
    page="(global)",
    page_file="(reused)",
    section="Treatment room wide",
    current_state="text only",
    photo_subject="Empty treatment room with reclined bed, device draped, accent lamp — 'private and clean' tone.",
    composition="wide",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="warm clinical neutral",
    must_include="treatment bed, draped device, branded touches",
    must_avoid="biohazard bins prominent, used disposables",
    consent_required="No",
    suggested_filename="src/img/photos/rooms/treatment-room-wide.webp",
    count="3 (different rooms)",
    notes="Reusable across procedure LPs.",
)
add(
    priority="P1",
    page="(global)",
    page_file="(reused)",
    section="Consultation room wide",
    current_state="text only",
    photo_subject="Consultation room wide — desk with iPad, two chairs, warm lamp, anatomy print framed on wall.",
    composition="wide",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="warm daylight + tungsten lamp",
    must_include="desk, two chairs, anatomy/aesthetic art",
    must_avoid="patient charts visible, prices on screen",
    consent_required="No",
    suggested_filename="src/img/photos/rooms/consultation-room-wide.webp",
    count="2",
    notes="Reused on /booking, /design-method, /decision-protection.",
)
add(
    priority="P1",
    page="/gallery.html",
    page_file="src/fragments/en/gallery.html",
    section="Hero (line 55)",
    current_state="text only on dark",
    photo_subject="Process imagery — physician's hand placing a printed plan card next to closed before/after photo envelopes — implies documentation without revealing patients.",
    composition="overhead",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="soft moody premium",
    must_include="plan card, archive folder/envelope, gloved hand",
    must_avoid="any actual before/after image (separate consent track)",
    consent_required="No",
    suggested_filename="src/img/photos/gallery/hero-documentation-process.webp",
    count="2",
    notes="Real before/after photos require their own patient release process — this row covers chrome only.",
)

# ============================================================
# P2 — nice-to-have
# ============================================================

add(
    priority="P2",
    page="/booking-manage.html",
    page_file="src/fragments/en/booking-manage.html",
    section="Page header",
    current_state="text only / form chrome",
    photo_subject="Calendar and pen on neutral surface — manage-your-booking visual cue.",
    composition="overhead",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="soft daylight",
    must_include="undated calendar, pen, branded card",
    must_avoid="real patient name on calendar",
    consent_required="No",
    suggested_filename="src/img/photos/booking-manage/header-calendar.webp",
    count="1",
    notes="Optional — useful if we later add a header strip image.",
)
add(
    priority="P2",
    page="/consult.html",
    page_file="src/fragments/en/consult.html",
    section="Hero",
    current_state="text only",
    photo_subject="Physician at desk with phone in hand mid-call (back to camera) — 'we'll get back to you' reassurance.",
    composition="medium",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="warm daylight",
    must_include="white-coat back, branded notepad, calm desk",
    must_avoid="visible phone screen with names",
    consent_required="Yes (staff release)",
    suggested_filename="src/img/photos/consult/hero-followup.webp",
    count="1",
    notes="Lower priority — current consult flow is mostly form.",
)
add(
    priority="P2",
    page="(blog)",
    page_file="src/blog/en/*.md (generic pattern)",
    section="Blog hero generic pool",
    current_state="text only",
    photo_subject="Three reusable detail crops: (a) anatomy print + marker, (b) printed Tune brochure on stone, (c) macro of fresh florals at reception. To be assigned to blog posts based on topic.",
    composition="close-up",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="warm daylight",
    must_include="brand-consistent detail, no patient",
    must_avoid="topic-specific elements (kept generic)",
    consent_required="No",
    suggested_filename="src/img/photos/blog/blog-hero-generic-{01-03}.webp",
    count="3 generic crops",
    notes="Editor maps each blog post to one of the 3 crops; topic-specific posts can be upgraded later.",
)
add(
    priority="P2",
    page="(global footer)",
    page_file="(reused)",
    section="Footer / contact strip",
    current_state="text only",
    photo_subject="Wide of building exterior at dusk with warm interior glow — quiet brand close.",
    composition="wide",
    orientation=LANDSCAPE_HERO,
    min_resolution=DEF_HERO_RES,
    mood_lighting="dusk premium, warm window glow",
    must_include="building, lit windows, sense of closure",
    must_avoid="people in frame, headlights blowing exposure",
    consent_required="No",
    suggested_filename="src/img/photos/global/footer-dusk.webp",
    count="2",
    notes="Captures end-of-day brand mood.",
)
add(
    priority="P2",
    page="(global)",
    page_file="(reused)",
    section="Concierge candid (English-speaking)",
    current_state="text only",
    photo_subject="English-speaking concierge laughing with stand-in international patient — natural candid energy.",
    composition="medium",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="warm daylight",
    must_include="concierge in uniform, English signage subtle",
    must_avoid="forced smile, stiff pose",
    consent_required="Yes (staff + stand-in release)",
    suggested_filename="src/img/photos/concierge/english-support-candid.webp",
    count="3",
    notes="Use sparingly; great for FAQ + international guide.",
)
add(
    priority="P2",
    page="(global)",
    page_file="(reused)",
    section="Skincare / aftercare product line",
    current_state="text only",
    photo_subject="Aftercare product set on counter — sun protection, calming serum, sheet mask.",
    composition="overhead",
    orientation=SQUARE,
    min_resolution=DEF_SQUARE_RES,
    mood_lighting="soft daylight",
    must_include="aftercare set, branded card",
    must_avoid="competitor brands without permission",
    consent_required="No",
    suggested_filename="src/img/photos/aftercare/product-set.webp",
    count="2",
    notes="Useful for blog hero pool and recovery posts.",
)
add(
    priority="P2",
    page="(global)",
    page_file="(reused)",
    section="View from clinic window",
    current_state="text only",
    photo_subject="View through 5F window of Apgujeong street — soft focus on indoor plant in foreground.",
    composition="medium",
    orientation=LANDSCAPE_3_2,
    min_resolution=DEF_LARGE_RES,
    mood_lighting="bright daylight + indoor warm",
    must_include="window framing, district view, plant detail",
    must_avoid="patient at window, identifiable signage outside",
    consent_required="No",
    suggested_filename="src/img/photos/location/window-view-apgujeong.webp",
    count="2",
    notes="Atmospheric; reuses across blog and global imagery.",
)

# ---------------------------------------------------------------------------
# Assign sequential IDs grouped by priority (preserve insertion order within
# each priority bucket since that order roughly tracks page-then-section).
# ---------------------------------------------------------------------------
priority_order = ["P0", "P1", "P2"]
counters = defaultdict(int)
# First pass: stable index per priority
sorted_rows: list[dict] = []
for pr in priority_order:
    for r in ROWS:
        if r["priority"] == pr:
            counters[pr] += 1
            r["id"] = f"{pr}-{counters[pr]:03d}"
            sorted_rows.append(r)

# ---------------------------------------------------------------------------
# Sheet 3: shoot day blocks. Rows are resolved by substring match against
# `suggested_filename` so the plan stays accurate even if rows are reordered.
# ---------------------------------------------------------------------------
def find_id(keyword: str) -> str:
    for r in sorted_rows:
        if keyword in r["suggested_filename"]:
            return r["id"]
    raise SystemExit(f"shoot-day-plan: no row matches keyword '{keyword}'")


def ids(*keywords: str) -> list[str]:
    return [find_id(k) for k in keywords]


BLOCKS = [
    {
        "block": "Block A - Exterior + district context",
        "time_window": "09:30 - 11:00 (morning daylight)",
        "lighting": "Natural daylight, no flash; consider polarizer for storefront glass",
        "rows": ids(
            "home/hero-clinic-exterior-golden-hour",
            "location/building-exterior-wide",
            "location/apgujeong-rodeo-street",
            "apgujeong/hero-district-clinic",
            "location/window-view-apgujeong",
            "global/footer-dusk",
        ),
        "duration_min": 90,
        "kit": "24-70mm, 16-35mm wide, polarizer, ND8, no tripod (handheld OK at this hour)",
    },
    {
        "block": "Block B - 5F approach + reception + waiting",
        "time_window": "11:00 - 12:30 (pre-opening for empty rooms)",
        "lighting": "Mixed daylight + warm tungsten; bring small bounce + softbox",
        "rows": ids(
            "home/visit-us-5f-arrival",
            "location/reception-waiting-wide",
            "home/inside-clinic-reception",
            "menu/hero-printed-menu",
            "concierge/english-support-candid",
            "aftercare/product-set",
        ),
        "duration_min": 90,
        "kit": "24-70mm, 35mm prime, tripod, 1 softbox, white bounce",
    },
    {
        "block": "Block C - Consultation rooms + treatment rooms",
        "time_window": "13:00 - 15:00 (skip lunch hour overlap with patients)",
        "lighting": "Daylight + tungsten; ensure no patient handover during shoot",
        "rows": ids(
            "rooms/treatment-room-wide",
            "rooms/consultation-room-wide",
            "gallery/hero-documentation-process",
            "booking/hero-consultation-desk-overhead",
            "decision-protection/hero-planning-desk",
            "design-method/hero-vector-mapping",
            "design-method/workflow-step",
            "how-to-choose/hero-indication-mapping",
        ),
        "duration_min": 120,
        "kit": "Tripod for overhead rig, 35mm/50mm primes, color checker, mannequin head if available",
    },
    {
        "block": "Block D - Equipment close-ups",
        "time_window": "15:00 - 16:30 (between booked slots; coordinate with nurse)",
        "lighting": "Clean clinical neutral; one warm rim light",
        "rows": ids(
            "metacell/equipment-prp-centrifuge",
            "metacell/build-",
            "signature-lifting/hero-ultherapy-handpiece",
            "structural-reset/hero-thermage-tip",
            "collagen-builder/hero-skin-booster-tray",
            "faq/hero-concierge-english-support",
            "korean-lifting-guide/hero-devices-lineup",
            "ultherapy-vs-thermage/hero-side-by-side",
            "dermal-fillers/hero-filler-tray",
            "skin-boosters/hero-vial-tray",
            "hair-loss/hero-scalp-treatment",
            "equipment/ultherapy-handpiece",
            "equipment/thermage-flx-tip",
            "equipment/onda-lifting",
        ),
        "duration_min": 90,
        "kit": "Macro 100mm, 50mm, tripod, two softboxes, gel filters, sterile pad set, color checker",
    },
    {
        "block": "Block D-2 - PBM + product still lifes",
        "time_window": "16:30 - 17:15",
        "lighting": "Same setup as Block D",
        "rows": ids(
            "equipment/pbm-device",
            "equipment/juvelook-rejuran-still",
        ),
        "duration_min": 45,
        "kit": "(continuation of Block D)",
    },
    {
        "block": "Block E - Physician portraits",
        "time_window": "17:30 - 19:00 (after patient hours; quiet space)",
        "lighting": "Window daylight + reflector; backup softbox if cloudy",
        "rows": ids(
            "doctors/dr-cha-portrait-formal",
            "doctors/dr-cha-consultation",
            "doctors/dr-ju-portrait-formal",
            "doctors/dr-ju-international-consult",
            "filler-chamaka-se/hero-design-mapping",
            "international-guide/hero-trip-planning-desk",
            "booking/success-welcome-florals",
        ),
        "duration_min": 90,
        "kit": "85mm portrait prime, 50mm, large reflector, softbox, lapel/clip mics if also recording video B-roll",
    },
    {
        "block": "Block F - Comfort + amenity (post-hours)",
        "time_window": "19:00 - 19:45 (lights warm, lounge calm)",
        "lighting": "Warm tungsten + soft fill",
        "rows": ids(
            "comfort/iv-sedation-chair",
            "comfort/recovery-lounge",
        ),
        "duration_min": 45,
        "kit": "24-70mm, tripod, 1 softbox",
    },
    {
        "block": "Block G - Generic blog hero pool + leftover details",
        "time_window": "Backstop slot — flex into other blocks if time allows",
        "lighting": "As needed",
        "rows": ids(
            "blog/blog-hero-generic",
            "booking-manage/header-calendar",
            "consult/hero-followup",
        ),
        "duration_min": 30,
        "kit": "(uses leftover setups)",
    },
]

# ---------------------------------------------------------------------------
# Page-summary notes (Sheet 2)
# ---------------------------------------------------------------------------
PAGE_NOTES = {
    "/": "Home - highest reuse. Pair Block A exterior + Block E portraits to land 5/6 ad launch.",
    "/booking.html": "Pair with Dr. Cha environmental from Block E; one overhead desk shot covers hero + success.",
    "/metacell-protocol.html": "Flagship LP - block off Dr. Cha or Dr. Ju gloved hands at start of Block D for the 3 build crops.",
    "(global)": "Reused across many pages - shoot once, wire everywhere.",
    "(equipment library)": "Coordinate with nurse during Block D to ensure devices are not in patient use.",
    "(blog)": "P2 - generic crops only; don't slow other blocks for these.",
    "(global footer)": "Revisit at dusk (18:00) for the warm-glow exterior variant.",
}

# ---------------------------------------------------------------------------
# Build the workbook
# ---------------------------------------------------------------------------
def build_xlsx() -> None:
    wb = Workbook()

    # --- Sheet 1: Shotlist ---
    ws = wb.active
    ws.title = "Shotlist"
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F172A")
    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, (name, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = border

    for r in sorted_rows:
        ws.append([r.get(c[0], "") for c in COLUMNS])

    # column widths + wrapping
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width

    last_row = ws.max_row
    last_col = len(COLUMNS)
    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            col_name = COLUMNS[cell.column - 1][0]
            cell.alignment = Alignment(
                vertical="top",
                horizontal="left",
                wrap_text=col_name in WRAP_COLS,
            )
            cell.border = border

    # priority fill on column 'priority' (column 2) and apply to whole row light tint
    pri_col_idx = headers.index("priority") + 1
    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
        pri = row[pri_col_idx - 1].value
        fill = PRIORITY_FILL.get(pri)
        if fill:
            row[pri_col_idx - 1].fill = fill
            row[pri_col_idx - 1].font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.row_dimensions[1].height = 30

    # --- Sheet 2: Summary by Page ---
    ws2 = wb.create_sheet("Summary by Page")
    ws2.append(["page", "page_file", "P0", "P1", "P2", "total", "shoot_day_notes"])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        c.border = border

    bucket: "OrderedDict[tuple[str,str], dict]" = OrderedDict()
    for r in sorted_rows:
        key = (r["page"], r["page_file"])
        b = bucket.setdefault(
            key,
            {"P0": 0, "P1": 0, "P2": 0, "total": 0},
        )
        b[r["priority"]] += 1
        b["total"] += 1

    for (page, page_file), counts in bucket.items():
        note = PAGE_NOTES.get(page, "")
        ws2.append(
            [page, page_file, counts["P0"], counts["P1"], counts["P2"], counts["total"], note]
        )

    widths_summary = [22, 50, 6, 6, 6, 8, 60]
    for i, w in enumerate(widths_summary, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
            cell.border = border
    ws2.freeze_panes = "A2"

    # totals row
    p0 = sum(b["P0"] for b in bucket.values())
    p1 = sum(b["P1"] for b in bucket.values())
    p2 = sum(b["P2"] for b in bucket.values())
    total = p0 + p1 + p2
    ws2.append(["TOTAL", "", p0, p1, p2, total, ""])
    for cell in ws2[ws2.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F1F5F9")
        cell.border = border

    # --- Sheet 3: Shoot Day Plan ---
    ws3 = wb.create_sheet("Shoot Day Plan")
    ws3.append(["block", "time_window", "duration_min", "kit_lighting", "covered_row_ids", "row_count"])
    for c in ws3[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        c.border = border

    total_minutes = 0
    total_rows_covered = 0
    for blk in BLOCKS:
        ids = blk["rows"]
        ws3.append(
            [
                blk["block"],
                blk["time_window"],
                blk["duration_min"],
                f"Lighting: {blk['lighting']}\nKit: {blk['kit']}",
                ", ".join(ids),
                len(ids),
            ]
        )
        total_minutes += blk["duration_min"]
        total_rows_covered += len(ids)

    widths_plan = [40, 30, 14, 60, 60, 12]
    for i, w in enumerate(widths_plan, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
            cell.border = border
    ws3.freeze_panes = "A2"

    ws3.append([])
    ws3.append([f"Total scheduled minutes", "", total_minutes, "", "", total_rows_covered])
    for cell in ws3[ws3.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F1F5F9")

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}  ({len(sorted_rows)} rows)")


# ---------------------------------------------------------------------------
# Build the markdown file
# ---------------------------------------------------------------------------
INTRO = """# Photo Shotlist - Tune Clinic Apgujeong - 2026-04-27

**Hand to staff before the shoot.** This file mirrors `photo-shotlist-2026-04-27.xlsx` (Sheet 1) so it's skimmable in Cursor and Drive.

How to use it:

1. **Shoot P0 first.** P0 rows block the 5/6 ad launch — if any P0 is missed, reschedule before any P1 work.
2. **Hand-off to dev.** Drop final assets into `src/img/photos/{slug}/{filename}.webp` exactly as listed under `suggested_filename`. We'll wire them into the page sections in a follow-up commit.
3. **Patient identifiable shots need signed releases.** Coordinate with reception. Stand-ins (back-of-head, hands only) are fine without a full release — see `consent_required` per row.
4. **Final delivery format:** `.webp`, sRGB, 80% quality. Keep RAW/HEIC originals in the shared Drive folder (`/photo-shoot-2026-04-27/raw/`).
5. **Lighting consistency matters more than perfection.** When a section calls for matched cards (e.g. Metacell three builds, Design Method workflow steps), shoot all variants in the same lighting setup back-to-back.
6. **Compliance reminders:** equal-size brand presence in product trays; no language implying that PRP/PBM is "stem cell" therapy; obscure serial numbers and lot labels.

See Sheet 3 ("Shoot Day Plan") in the xlsx for time-blocked groupings (Blocks A through G).

---

## Shotlist
"""

MD_COLS = [c[0] for c in COLUMNS]


def md_escape(v) -> str:
    s = str(v) if v is not None else ""
    return s.replace("|", "\\|").replace("\n", " ")


def build_md() -> None:
    out = [INTRO]
    out.append("| " + " | ".join(MD_COLS) + " |")
    out.append("|" + "|".join(["---"] * len(MD_COLS)) + "|")
    for r in sorted_rows:
        out.append("| " + " | ".join(md_escape(r.get(c, "")) for c in MD_COLS) + " |")

    # priority counts
    p0 = sum(1 for r in sorted_rows if r["priority"] == "P0")
    p1 = sum(1 for r in sorted_rows if r["priority"] == "P1")
    p2 = sum(1 for r in sorted_rows if r["priority"] == "P2")
    out.append("")
    out.append(f"**Totals:** {len(sorted_rows)} rows ({p0} P0, {p1} P1, {p2} P2)")
    out.append("")

    os.makedirs(os.path.dirname(OUT_MD), exist_ok=True)
    with open(OUT_MD, "w") as f:
        f.write("\n".join(out) + "\n")
    print(f"wrote {OUT_MD}")


if __name__ == "__main__":
    build_xlsx()
    build_md()
