#!/usr/bin/env python3
"""
심플 촬영 안내 엑셀 (직원 전달용).
`photo-shotlist-ko.json`과 동일한 52행을, '무엇을 / 어떻게'만 짧게 정리합니다.
"""
from __future__ import annotations

import json
import os
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
KO_PATH = os.path.join(REPO, "scripts", "photo-shotlist-ko.json")
OUT_XLSX = os.path.join(REPO, "docs", "photo-shotlist-simple-2026-04-27.xlsx")


def sort_by_priority(rows: list[dict]) -> list[dict]:
    order = ["P0", "P1", "P2"]
    counters: dict[str, int] = defaultdict(int)
    out: list[dict] = []
    for pr in order:
        for r in rows:
            if r.get("priority") == pr:
                counters[pr] += 1
                out.append(dict(r))
    return out


def build() -> None:
    with open(KO_PATH, encoding="utf-8") as f:
        rows: list[dict] = json.load(f)
    rows = sort_by_priority(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "촬영목록"

    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")

    # 안내 1행 (병합)
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = (
        "홈페이지에 올릴 사진 촬영 리스트입니다. "
        "P0(최우선) → P1 → P2 순으로 진행해 주세요. "
        "얼굴이 나오면 사전에 동의가 필요합니다."
    )
    c.font = Font(size=11)
    c.alignment = Alignment(wrap_text=True, vertical="center")

    headers = ["번호", "순서", "이런 사진을 찍어 주세요", "이렇게 찍어 주세요"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for i, r in enumerate(rows, start=1):
        pr = r.get("priority", "")
        what = f"{r.get('photo_subject', '')}".strip()
        sec = (r.get("section") or "").strip()
        if sec:
            what = f"({sec})\n{what}"

        consent = r.get("consent_required", "")
        consent_line = "" if consent in ("", "없음") else f" · 동의 {consent}"

        how = "\n".join(
            [
                f"① {r.get('composition', '')} / {r.get('orientation', '')} · {r.get('mood_lighting', '')}",
                f"② {r.get('count', '')}{consent_line}",
                f"③ 꼭 담기: {r.get('must_include', '')}",
                f"④ 피하기: {r.get('must_avoid', '')}",
            ]
        )
        notes = (r.get("notes") or "").strip()
        if notes:
            how += f"\n⑤ {notes}"

        row_idx = i + 2
        ws.cell(row=row_idx, column=1, value=i)
        ws.cell(row=row_idx, column=2, value=pr)
        ws.cell(row=row_idx, column=3, value=what)
        ws.cell(row=row_idx, column=4, value=how)
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).alignment = wrap

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 52
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    for row in range(3, len(rows) + 3):
        pri = ws.cell(row=row, column=2).value
        fill = None
        if pri == "P0":
            fill = PatternFill("solid", fgColor="FEE2E2")
        elif pri == "P1":
            fill = PatternFill("solid", fgColor="FFEDD5")
        elif pri == "P2":
            fill = PatternFill("solid", fgColor="E2E8F0")
        if fill:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX} ({len(rows)} rows)")


if __name__ == "__main__":
    build()
